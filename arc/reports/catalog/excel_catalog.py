import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from settings import settings
from arc.settings import settings as py_settings
from arc.contrib.steps.api import api_keywords
from arc.contrib.steps.general import funcional_keywords, datas_keywords, ftp_keywords
from arc.contrib.steps.web import web_keywords
from arc.reports.catalog.pydoc_formatter import get_pydoc_info
from arc.reports.catalog.user_step import get_list_user_steps

PATH_HOME = py_settings.OUTPUT_PATH + os.sep


class Catalog:
    workbook = None
    functions_list = []
    options_talos = {}
    all_default_function_config = {}
    user_functions_list = []
    all_user_function_config = {}
    step_verbs = ["step", "given", "when", "then", "and"]

    def __init__(self):
        self.workbook = Workbook()
        self.get_config()
        self.get_user_function_datas()
        self.write_function_user_data_in_excel()
        self.get_default_function_datas()
        self.write_function_default_data_in_excel()
        self.wordfinder()
        self.save_excel()

    def set_sheet_title(self, title_name):
        worksheet = self.workbook[title_name]
        worksheet['A1'] = title_name
        worksheet.merge_cells('A1:F1')
        font = Font(color="FF0000", bold=True, sz=18)
        alignment = Alignment(horizontal="center", vertical="center")
        a1 = worksheet['A1']
        a1.font = font
        a1.alignment = alignment

    def set_headers_title(self, sheet):
        worksheet = self.workbook[sheet]
        font = Font(bold=True, sz=12, color="FF0000")
        alignment = Alignment(horizontal="center", vertical="center")
        worksheet['A1'] = "Type"
        worksheet['A1'].font = font
        worksheet['A1'].alignment = alignment
        worksheet['B1'] = "Step"
        worksheet['B1'].font = font
        worksheet['B1'].alignment = alignment
        worksheet['C1'] = "Step Name"
        worksheet['C1'].font = font
        worksheet['C1'].alignment = alignment
        worksheet['D1'] = "Information"
        worksheet['D1'].font = font
        worksheet['D1'].alignment = alignment
        worksheet['E1'] = "Params"
        worksheet['E1'].font = font
        worksheet['E1'].alignment = alignment
        worksheet['F1'] = "Step Example"
        worksheet['F1'].font = font
        worksheet['F1'].alignment = alignment
        worksheet['G1'] = "File Path"
        worksheet['G1'].font = font
        worksheet['G1'].alignment = alignment

    @staticmethod
    def set_column_width(worksheet):
        ws = worksheet
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2

    @staticmethod
    def set_column_custom_width(worksheet, column, width):
        worksheet.column_dimensions[column].width = width

    def get_user_function_datas(self):
        if self.options_talos["user_steps"]:
            self.user_functions_list = get_list_user_steps()

            titles = []
            for lista in self.user_functions_list:
                for diccionarios in lista:
                    if diccionarios["Function Name"] not in titles:
                        titles.append(diccionarios["Function Name"])

                for title in titles:
                    if "'Function Name': '" + title + "'" in str(lista):
                        self.all_user_function_config[title] = {"Title": title, "Datas": lista}

    def write_function_user_data_in_excel(self):
        font = Font(bold=True, sz=12)
        for function_key in self.all_user_function_config.keys():
            title = ""
            all_fuctions = []
            for each_key in self.all_user_function_config[function_key].keys():
                if each_key == "Title":
                    title = self.all_user_function_config[function_key][each_key]

                if each_key == "Datas":
                    all_fuctions = self.all_user_function_config[function_key][each_key]

                if "steps" not in str(title).lower() or "step" not in str(title).lower():
                    title = title + " Steps"
                self.create_sheet(title)

            worksheet = self.workbook[title]
            self.set_headers_title(title)
            cont = 2
            for list_with_dict in all_fuctions:
                function_tag = list_with_dict["Tags"]
                verb_step = list_with_dict["Verb Step"]
                step = list_with_dict["Step"]
                py_doc = list_with_dict["Py Doc"]
                example = list_with_dict["Example"]
                params = list_with_dict["Params"]
                file_paths = list_with_dict["Function Path"]
                if str(verb_step).lower() in self.step_verbs:
                    if function_tag is None: function_tag = "None"
                    worksheet['A' + str(cont)] = function_tag
                    worksheet['A' + str(cont)].font = font
                    worksheet['A' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    worksheet['B' + str(cont)] = verb_step
                    worksheet['B' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    worksheet['C' + str(cont)] = str(step).replace("(u'", "").replace("')", "").replace("(\"", "")
                    worksheet['C' + str(cont)].alignment = Alignment(vertical="center")

                    worksheet['D' + str(cont)].alignment = Alignment(vertical="center", wrapText=True)
                    text_doc = ""
                    for doc in py_doc:
                        text_doc = text_doc + doc + "\n"
                    worksheet['D' + str(cont)] = text_doc[:-1]

                    worksheet['E' + str(cont)].alignment = Alignment(vertical="center", horizontal="center",
                                                                     wrapText=True)
                    text_param = ""
                    for param in params:
                        text_param = text_param + param + "\n"
                    worksheet['E' + str(cont)] = text_param[:-1]

                    example_list = str(example).split("\n")
                    worksheet['F' + str(cont)].alignment = Alignment(vertical="center", horizontal="center",
                                                                     wrapText=True)
                    text_example = ""
                    for example_lines in example_list:
                        text_example = text_example + example_lines + "\n"
                    worksheet['F' + str(cont)] = text_example[:-1]

                    font_file = Font(bold=False, sz=12)
                    worksheet['G' + str(cont)] = file_paths
                    worksheet['G' + str(cont)].font = font_file
                    worksheet['G' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    cont += 1

            self.set_column_width(worksheet)
            self.set_column_custom_width(worksheet, "D", 40)
            self.set_column_custom_width(worksheet, "F", 40)

    def get_default_function_datas(self):
        if self.options_talos["default_api"]:
            self.all_default_function_config["default_api"] = {"Title": "Api Default Steps",
                                                              "Datas": get_pydoc_info(api_keywords)}
        if self.options_talos["default_web"]:
            self.all_default_function_config["default_web"] = {"Title": "Web Default Steps",
                                                              "Datas": get_pydoc_info(web_keywords)}
            pass
        if self.options_talos["default_funcional"]:
            self.all_default_function_config["funcional_keywords"] = {"Title": "Funcional Default Steps",
                                                                     "Datas": get_pydoc_info(funcional_keywords)}
        if self.options_talos["default_datas"]:
            self.all_default_function_config["default_datas"] = {"Title": "Datas Default Steps",
                                                                "Datas": get_pydoc_info(datas_keywords)}

        if self.options_talos["default_ftp"]:
            self.all_default_function_config["default_ftp"] = {"Title": "FTP Default Steps",
                                                              "Datas": get_pydoc_info(ftp_keywords)}

    def write_function_default_data_in_excel(self):
        font = Font(bold=True, sz=12)
        for function_key in self.all_default_function_config.keys():
            title = ""
            all_fuctions = []
            for each_key in self.all_default_function_config[function_key].keys():
                if each_key == "Title":
                    title = self.all_default_function_config[function_key][each_key]
                if each_key == "Datas":
                    all_fuctions = self.all_default_function_config[function_key][each_key]
                self.create_sheet(title)

            worksheet = self.workbook[title]
            self.set_headers_title(title)
            cont = 2
            for list_with_dict in all_fuctions:
                function_tag = list_with_dict["Tags"]
                verb_step = list_with_dict["Verb Step"]
                step = list_with_dict["Step"]
                py_doc = list_with_dict["Py Doc"]
                example = list_with_dict["Example"]
                params = list_with_dict["Params"]
                file_paths = list_with_dict["Function Path"]
                if str(verb_step).lower() in self.step_verbs:
                    if function_tag is None: function_tag = "None"
                    worksheet['A' + str(cont)] = function_tag
                    worksheet['A' + str(cont)].font = font
                    worksheet['A' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    worksheet['B' + str(cont)] = verb_step
                    worksheet['B' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    worksheet['C' + str(cont)] = str(step).replace("(u'", "").replace("')", "").replace("(\"", "")
                    worksheet['C' + str(cont)].alignment = Alignment(vertical="center")

                    worksheet['D' + str(cont)].alignment = Alignment(vertical="center", wrapText=True)
                    text_doc = ""
                    for doc in py_doc:
                        text_doc = text_doc + doc + "\n"
                    worksheet['D' + str(cont)] = text_doc[:-1]

                    worksheet['E' + str(cont)].alignment = Alignment(vertical="center", horizontal="center",
                                                                     wrapText=True)
                    text_param = ""
                    for param in params:
                        text_param = text_param + param + "\n"
                    worksheet['E' + str(cont)] = text_param[:-1]

                    example_list = str(example).split("\n")
                    worksheet['F' + str(cont)].alignment = Alignment(vertical="center", horizontal="center",
                                                                     wrapText=True)
                    text_example = ""
                    for example_lines in example_list:
                        text_example = text_example + example_lines + "\n"
                    worksheet['F' + str(cont)] = text_example[:-1]

                    font_file = Font(bold=False, sz=12)
                    worksheet['G' + str(cont)] = file_paths
                    worksheet['G' + str(cont)].font = font_file
                    worksheet['G' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    cont += 1

            self.set_column_width(worksheet)
            self.set_column_custom_width(worksheet, "D", 40)
            self.set_column_custom_width(worksheet, "F", 40)

    def create_sheet(self, sheet):
        if sheet not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet)

        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

        self.set_headers_title(sheet)

    def get_config(self):
        self.options_talos = {
            "default_api": settings.PYTALOS_CATALOG['steps']['default_api'],
            "default_web": settings.PYTALOS_CATALOG['steps']['default_web'],
            "default_funcional": settings.PYTALOS_CATALOG['steps']['default_functional'],
            "default_datas": settings.PYTALOS_CATALOG['steps']['default_data'],
            "default_ftp": settings.PYTALOS_CATALOG['steps']['default_ftp'],
            "user_steps": settings.PYTALOS_CATALOG['steps']['user_steps']
        }

    def wordfinder(self):
        values = ["STEP", "GIVEN", "WHEN", "THEN", "AND", "None"]

        font_verb_step = Font(color="5C0303", bold=True, sz=10)
        font_verb_given = Font(color="030E5C", bold=True, sz=10)
        font_verb_when = Font(color="035C13", bold=True, sz=10)
        font_verb_then = Font(color="5C0353", bold=True, sz=10)
        font_verb_and = Font(color="03595C", bold=True, sz=10)
        font_verb_none = Font(color="ED0606", bold=True, sz=10)

        for sheet in self.workbook.sheetnames:
            worksheet = self.workbook[sheet]
            for i in range(2, worksheet.max_row + 1):
                for j in range(1, worksheet.max_column + 1):
                    for value in values:
                        if str(value) == str(worksheet.cell(row=i, column=j).value):
                            if value == "STEP":
                                worksheet.cell(row=i, column=j).font = font_verb_step
                            if value == "GIVEN":
                                worksheet.cell(row=i, column=j).font = font_verb_given
                            if value == "WHEN":
                                worksheet.cell(row=i, column=j).font = font_verb_when
                            if value == "THEN":
                                worksheet.cell(row=i, column=j).font = font_verb_then
                            if value == "AND":
                                worksheet.cell(row=i, column=j).font = font_verb_and
                            if value == "None":
                                worksheet.cell(row=i, column=j).font = font_verb_none

    def save_excel(self):
        output_path_catalog = str(PATH_HOME
                                  + settings.PYTALOS_CATALOG['excel_file_name'] + ".xlsx").replace("\\", "/")
        self.workbook.save(filename=output_path_catalog)